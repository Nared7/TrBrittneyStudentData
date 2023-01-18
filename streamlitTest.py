# import pandas as pd
import pandas as pd
import streamlit as st
import openpyxl
# import matplotlib.pyplot as plt
from string import capwords


Name = st.text_input('Which student are you looking for?')

wb = openpyxl.load_workbook(r'studentCost.xlsx', data_only=True)
janInvoiceWB = openpyxl.load_workbook(r'InvoiceList_Jan2023.xlsx')


def findStudent(excelWB, studentName):
    ssDatabaseName = ''
    studentClassName = ''
    studentClass = ''
    studentSheetRow = ''

    engClassPrice = ''
    engClassDiscount = ''
    engClassNetPrice = ''

    grammarClassPrice = ''
    grammarClassDiscount = ''
    grammarClassNetPrice = ''

    bookFee = ''

    totalCost = ''

    receiptNote = ''

    firstPerson = 0
    stName = ''

    for cnt, sheet in enumerate(wb.sheetnames):
        ws = excelWB[sheet]
        if firstPerson:
            firstPerson = 0
            break
        for row in range(2,13):
            stName = ws['B'+str(row)].value

            try:
                stName = stName.lower()
            except:
                pass

            if stName == studentName.lower():
                ssDatabaseName = stName
                studentClass = ws
                studentClassName = sheet
                if sheet == 'YL Starter' or sheet == 'PreYL' or sheet == 'Teen' or sheet == 'Upper Intermediate 1' or \
                    sheet == 'Intermediate 1' or sheet == 'Intermediate 2' or sheet == 'Pre-Intermediate 1' or sheet == 'Pre-Intermediate 2':
                    engClassPrice = str("{:,}".format(studentClass['C' + str(row)].value)) + ' kyats'
                    engClassDiscount = str((studentClass['D' + str(row)].value) * 100) + ' %'
                    engClassNetPrice = str("{:,}".format(studentClass['E' + str(row)].value)) + ' kyats'

                    grammarClassPrice = '-'
                    grammarClassDiscount = '-'
                    grammarClassNetPrice = '-'

                    bookPrice = studentClass['F' + str(row)].value

                    if bookPrice is None:
                        bookFee = '-     '
                    else:
                        bookFee = str("{:,}".format(studentClass['F' + str(row)].value))  + ' kyats'

                    totalCost = str("{:,}".format(studentClass['G' + str(row)].value)) + ' kyats'
                else:
                    engClassPrice = str("{:,}".format(studentClass['C' + str(row)].value)) + ' kyats'
                    engClassDiscount = str((studentClass['D' + str(row)].value) * 100) + ' %'
                    engClassNetPrice = str("{:,}".format(studentClass['E' + str(row)].value)) + ' kyats'

                    grammarClassPrice = str("{:,}".format(studentClass['F' + str(row)].value)) + ' kyats'
                    grammarClassDiscount = str((studentClass['G' + str(row)].value) * 100) + ' %'
                    grammarClassNetPrice = str("{:,}".format(studentClass['H' + str(row)].value)) + ' kyats'

                    bookPrice = studentClass['I' + str(row)].value

                    if bookPrice is None:
                        bookFee = '-     '
                    else:
                        bookFee = str("{:,}".format(studentClass['I' + str(row)].value)) + ' kyats'

                    totalCost = str("{:,}".format(studentClass['J' + str(row)].value)) + ' kyats'
                firstPerson = 1
                break # To get first Daniel
    if studentClass == '':
        ssDatabaseName = 'No such student'

    toReturnData = [ssDatabaseName, studentClassName, engClassPrice, engClassDiscount,engClassNetPrice,grammarClassPrice, grammarClassDiscount, grammarClassNetPrice, bookFee, totalCost]
    return toReturnData


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def findStudentInvoice(excelWB, studentName):
    stName = ''
    stClass = ''
    tranID = ''
    invoiceID = ''
    amount = ''

    for cnt, sheet in enumerate(excelWB.sheetnames):
        ws = excelWB[sheet]
        maxRow = get_maximum_rows(sheet_object=ws)
        for row in range(2, maxRow+1):
            stName = ws['A' + str(row)].value
            try:
                stName = stName.lower()
            except:
                pass
            if stName.lower() == studentName.lower():
                stClass = ws['B' + str(row)].value
                tranID = ws['C' + str(row)].value
                invoiceID = ws['D' + str(row)].value
                amount = ws['E' + str(row)].value
    toReturnList = [studentName, stClass, tranID, invoiceID, amount]
    return toReturnList


stData = [findStudent(wb, Name)]
stInvoiceData = [findStudentInvoice(janInvoiceWB, Name)]
stLabel = ['Name', 'Class', '4 Skill Class Fee', '4 Skill Discount', '4 Skill Net Fee', 'Grammar Class Fee', 'Grammar Discount', 'Grammar Net Fee', 'Book Fee', 'Total Cost']
stInvoiceLabel = ['Name', 'Class', 'Kpay ID', 'Invoice ID', 'Amount']
pdDataFrame = pd.DataFrame(stData, columns=stLabel)
pdInvoiceDataFrame = pd.DataFrame(stInvoiceData, columns=stInvoiceLabel)
st.table(pdDataFrame)
st.table(pdInvoiceDataFrame)

